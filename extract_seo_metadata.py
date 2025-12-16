import os
import re
import pandas as pd
from pathlib import Path

# Base directory
app_dir = Path(r"c:\Users\NIKHIL\Desktop\coreway-new\app")
base_url = "https://www.corewaysolution.com"  # Update with actual domain

# Function to extract metadata from a file (page.tsx or layout.tsx)
def extract_metadata(file_path):
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        metadata = {
            'title': '',
            'description': '',
            'keywords': ''
        }
        
        # Try to find metadata object
        # Pattern 1: export const metadata = { ... }
        metadata_match = re.search(r'export const metadata[:\s]*=\s*\{([^}]+(?:\{[^}]*\}[^}]*)*)\}', content, re.DOTALL)
        
        if metadata_match:
            metadata_content = metadata_match.group(1)
            
            # Extract title (handle both string and object formats)
            title_match = re.search(r'title:\s*["\']([^"\']+)["\']', metadata_content)
            if not title_match:
                # Try object format with default
                title_match = re.search(r'title:\s*\{[^}]*default:\s*["\']([^"\']+)["\']', metadata_content)
            if not title_match:
                # Try template literal
                title_match = re.search(r'title:\s*`([^`]+)`', metadata_content)
            if title_match:
                metadata['title'] = title_match.group(1)
            
            # Extract description
            desc_match = re.search(r'description:\s*["\']([^"\']+)["\']', metadata_content)
            if not desc_match:
                desc_match = re.search(r'description:\s*`([^`]+)`', metadata_content)
            if desc_match:
                metadata['description'] = desc_match.group(1)
            
            # Extract keywords
            keywords_match = re.search(r'keywords:\s*["\']([^"\']+)["\']', metadata_content)
            if not keywords_match:
                keywords_match = re.search(r'keywords:\s*`([^`]+)`', metadata_content)
            if keywords_match:
                metadata['keywords'] = keywords_match.group(1)
        
        # Also check for Head component (for client components)
        if not metadata['title']:
            head_title = re.search(r'<title>([^<]+)</title>', content)
            if head_title:
                metadata['title'] = head_title.group(1)
        
        if not metadata['description']:
            head_desc = re.search(r'<meta\s+name="description"\s+content="([^"]+)"', content)
            if head_desc:
                metadata['description'] = head_desc.group(1)
        
        if not metadata['keywords']:
            head_keywords = re.search(r'<meta\s+name="keywords"\s+content="([^"]+)"', content)
            if head_keywords:
                metadata['keywords'] = head_keywords.group(1)
        
        return {
            'title': metadata['title'].strip(),
            'description': metadata['description'].strip(),
            'keywords': metadata['keywords'].strip()
        }
    except Exception as e:
        print(f"Error reading {file_path}: {e}")
        return {
            'title': '',
            'description': '',
            'keywords': ''
        }

# Function to convert directory path to URL
def path_to_url(dir_path, app_dir, base_url):
    # Get relative path from app directory
    rel_path = dir_path.relative_to(app_dir)
    
    # Get path parts
    path_parts = list(rel_path.parts) if str(rel_path) != '.' else []
    
    # Handle dynamic routes
    url_parts = []
    for part in path_parts:
        if part.startswith('[') and part.endswith(']'):
            # Dynamic route - show as placeholder
            param_name = part[1:-1]
            url_parts.append(f'{{{param_name}}}')
        else:
            url_parts.append(part)
    
    # Build URL
    if not url_parts:
        return base_url + "/"
    else:
        return base_url + "/" + "/".join(url_parts)

# Main extraction
def extract_all_seo_data():
    data = []
    processed_dirs = set()
    
    # Find all page.tsx files
    page_files = list(app_dir.rglob("page.tsx"))
    
    print(f"Found {len(page_files)} pages")
    
    for page_file in sorted(page_files):
        page_dir = page_file.parent
        
        # Skip if already processed
        if page_dir in processed_dirs:
            continue
        processed_dirs.add(page_dir)
        
        # Skip admin pages
        if 'admin' in str(page_file):
            continue
        
        url = path_to_url(page_dir, app_dir, base_url)
        
        # Try to get metadata from layout.tsx first, then page.tsx
        layout_file = page_dir / "layout.tsx"
        metadata = {'title': '', 'description': '', 'keywords': ''}
        
        if layout_file.exists():
            metadata = extract_metadata(layout_file)
            source = "layout.tsx"
        
        # If no metadata in layout, try page.tsx
        if not metadata['title'] and not metadata['description']:
            metadata = extract_metadata(page_file)
            source = "page.tsx"
        
        data.append({
            'URL': url,
            'Title': metadata['title'],
            'Description': metadata['description'],
            'Keywords': metadata['keywords'],
            'File Path': str(page_file.relative_to(app_dir.parent))
        })
        
        print(f"Processed: {url}")
    
    return data

# Generate Excel file
def create_excel(data, output_file):
    df = pd.DataFrame(data)
    
    # Reorder columns
    df = df[['URL', 'Title', 'Description', 'Keywords', 'File Path']]
    
    # Sort by URL
    df = df.sort_values('URL')
    
    # Create Excel writer with formatting
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='SEO Metadata', index=False)
        
        # Get the worksheet
        worksheet = writer.sheets['SEO Metadata']
        
        # Auto-adjust column widths
        for idx, col in enumerate(df.columns):
            max_length = max(
                df[col].astype(str).apply(len).max(),
                len(col)
            )
            # Set a reasonable max width
            max_length = min(max_length, 80)
            # Excel column letters
            col_letter = chr(65 + idx)
            worksheet.column_dimensions[col_letter].width = max_length + 2
        
        # Make header row bold
        from openpyxl.styles import Font, PatternFill, Alignment
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Wrap text for description column
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=3, max_col=3):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Set row height for better readability
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            worksheet.row_dimensions[row[0].row].height = 30
    
    print(f"\nExcel file created: {output_file}")
    print(f"Total pages: {len(df)}")
    
    # Print summary statistics
    print(f"\nSummary:")
    print(f"  Pages with Title: {df['Title'].astype(bool).sum()}")
    print(f"  Pages with Description: {df['Description'].astype(bool).sum()}")
    print(f"  Pages with Keywords: {df['Keywords'].astype(bool).sum()}")
    print(f"  Pages missing metadata: {len(df) - df[['Title', 'Description']].any(axis=1).sum()}")

if __name__ == "__main__":
    print("Extracting SEO metadata from all pages...")
    print("=" * 60)
    seo_data = extract_all_seo_data()
    
    output_file = r"c:\Users\NIKHIL\Desktop\coreway-new\seo_metadata.xlsx"
    create_excel(seo_data, output_file)
    
    print("\n" + "=" * 60)
    print("✓ Done!")
    print(f"Excel file saved to: {output_file}")
